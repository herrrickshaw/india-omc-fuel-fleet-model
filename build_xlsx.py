#!/usr/bin/env python3
"""Build a formula-driven Excel version of the OMC retail profitability model.
Levers live on the Inputs sheet; all outputs are Excel formulas referencing
them, so changing a lever recalculates the whole workbook.
Run recalc.py (LibreOffice) afterwards to populate cached values.
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = Path(__file__).resolve().parent / "outputs" / "OMC_Retail_Profitability_Model.xlsx"
BLUE = Font(color="1F4E79", bold=True)      # editable input
GREEN = Font(color="217346")                # cross-sheet link
HEAD = Font(bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="305496")
TITLE = Font(bold=True, size=13)
NOTE = Font(italic=True, size=9, color="808080")
RS = u'#,##0'; RS2 = u'#,##0.00'; PCT = u'0.0%'
thin = Border(*[Side(style="thin", color="D9D9D9")] * 4)

wb = openpyxl.Workbook()

# ── Inputs ──────────────────────────────────────────────────────────────────
ip = wb.active; ip.title = "Inputs"; ip.sheet_view.showGridLines = False
ip["A1"] = "OMC Retail Profitability Model — Inputs (edit blue cells)"; ip["A1"].font = TITLE
rows = [
    ("Retail outlets (01.10.2025)", 99281, "count", "PPAC RR Table 6.6/6.7"),
    ("Outlets prior year (01.10.2024)", 91949, "count", "PPAC RR Table 6.4D"),
    ("Petrol (MS) consumption FY24-25", 40.0, "MMT", "PPAC RR Table 6.1"),
    ("Diesel (HSD) consumption FY24-25", 91.4, "MMT", "PPAC RR Table 6.1"),
    ("Petrol density", 0.74, "kg/L", "standard"),
    ("Diesel density", 0.83, "kg/L", "standard"),
    ("OMC marketing margin — petrol", 3.5, "Rs/L", "LEVER (dealer comm. excluded)"),
    ("OMC marketing margin — diesel", 2.5, "Rs/L", "LEVER"),
    ("Petrol demand growth (vehicle-driven)", 0.071, "%/yr", "RR 6.1 H1 YoY"),
    ("Diesel demand growth (vehicle-driven)", 0.029, "%/yr", "RR 6.1 H1 YoY"),
    ("Outlet network growth", 0.045, "%/yr", "LEVER"),
    ("E20 blend %", 0.20, "%", "achieved ~2025"),
    ("E20 mileage drop", 0.040, "%", "SIAM/ARAI central"),
    ("E25 blend %", 0.25, "%", "roadmap"),
    ("E25 mileage drop", 0.055, "%", "LEVER (LCV-scaled)"),
    ("E30 blend %", 0.30, "%", "roadmap"),
    ("E30 mileage drop", 0.070, "%", "LEVER (LCV-scaled)"),
]
ip.append(["", "", "", ""])
ip.append(["Parameter", "Value", "Unit", "Source / note"])
for c in ip[3]:
    c.font = HEAD; c.fill = HFILL
R = {}
for i, (name, val, unit, note) in enumerate(rows, start=4):
    ip.cell(i, 1, name)
    vc = ip.cell(i, 2, val); vc.font = BLUE
    vc.number_format = PCT if unit in ("%", "%/yr") else (RS2 if unit in ("Rs/L", "kg/L", "MMT") else RS)
    ip.cell(i, 3, unit); ip.cell(i, 4, note).font = NOTE
    R[name] = f"Inputs!$B${i}"
ip.column_dimensions["A"].width = 38; ip.column_dimensions["D"].width = 34

# helper refs
NRO, NRO0 = R["Retail outlets (01.10.2025)"], R["Outlets prior year (01.10.2024)"]
MS_MMT, HSD_MMT = R["Petrol (MS) consumption FY24-25"], R["Diesel (HSD) consumption FY24-25"]
DMS, DHSD = R["Petrol density"], R["Diesel density"]
MGN_MS, MGN_HSD = R["OMC marketing margin — petrol"], R["OMC marketing margin — diesel"]
GMS, GHSD, GRO = R["Petrol demand growth (vehicle-driven)"], R["Diesel demand growth (vehicle-driven)"], R["Outlet network growth"]

def bnL(mmt_ref, dens_ref):        # MMT -> bn L
    return f"({mmt_ref}*1000/{dens_ref})/1000"   # MMT*1e9/dens/1e9 = MMT*1000/dens/1000

# ── Base ────────────────────────────────────────────────────────────────────
b = wb.create_sheet("Base"); b.sheet_view.showGridLines = False
b["A1"] = "Base retail book (FY 2024-25 actual)"; b["A1"].font = TITLE
b.append([]); b.append(["Item", "Petrol (MS)", "Diesel (HSD)", "Total"])
for c in b[3]:
    c.font = HEAD; c.fill = HFILL
b.append(["Volume (bn L)", f"={bnL(MS_MMT,DMS)}", f"={bnL(HSD_MMT,DHSD)}", "=B4+C4"])
b.append(["OMC margin (Rs/L)", f"={MGN_MS}", f"={MGN_HSD}", ""])
b.append(["OMC retail gross (Rs cr/yr)", "=B4*1000000000*B5/10000000", "=C4*1000000000*C5/10000000", "=B6+C6"])
b.append(["Avg throughput/RO (KL/month)", f"=B4*1000000000/{NRO}/12/1000", f"=C4*1000000000/{NRO}/12/1000", ""])
for r in range(4, 8):
    for col in "BCD":
        cell = b[f"{col}{r}"]
        if r in (4, 8): cell.number_format = RS2
        elif r == 6: cell.number_format = RS
        else: cell.number_format = RS2
b.column_dimensions["A"].width = 30
for col in "BCD": b.column_dimensions[col].width = 15
BASE_TOT = "Base!$D$6"

# ── Scenarios ───────────────────────────────────────────────────────────────
s = wb.create_sheet("Scenarios"); s.sheet_view.showGridLines = False
s["A1"] = "Ethanol blend scenarios (petrol only; same distance driven)"; s["A1"].font = TITLE
s["A2"] = "L0 = E0-equivalent petrol volume (fixed distance) = base petrol blend × (1 − E20 drop)"; s["A2"].font = NOTE
hdr = ["Scenario", "Blend %", "Mileage drop", "Blend vol (bn L)", "Petrol MS (bn L)",
       "Ethanol (bn L)", "Extra vs E0 (bn L)", "Extra OMC income (Rs cr/yr)"]
s.append([]); s.append(hdr)
for c in s[4]:
    c.font = HEAD; c.fill = HFILL
L0 = f"({bnL(MS_MMT,DMS)})*(1-{R['E20 mileage drop']})"     # fixed-distance baseline
specs = [("E20", R["E20 blend %"], R["E20 mileage drop"]),
         ("E25", R["E25 blend %"], R["E25 mileage drop"]),
         ("E30", R["E30 blend %"], R["E30 mileage drop"])]
for i, (tag, pct, drop) in enumerate(specs):
    r = 5 + i
    blendv = f"({L0}/(1-{drop}))"
    s.cell(r, 1, tag)
    s.cell(r, 2, f"={pct}").number_format = PCT
    s.cell(r, 3, f"={drop}").number_format = PCT
    s.cell(r, 4, f"={blendv}").number_format = RS2
    s.cell(r, 5, f"=D{r}*(1-{pct})").number_format = RS2
    s.cell(r, 6, f"=D{r}*{pct}").number_format = RS2
    s.cell(r, 7, f"=D{r}-({L0})").number_format = RS2
    s.cell(r, 8, f"=G{r}*1000000000*{MGN_MS}/10000000").number_format = RS
s.column_dimensions["A"].width = 11
for col in "BCDEFGH": s.column_dimensions[col].width = 16

# ── BlendMix (E0/E20/E25/E30 in a ratio) ────────────────────────────────────
mx = wb.create_sheet("BlendMix"); mx.sheet_view.showGridLines = False
mx["A1"] = "Fuel-pool mix scenarios — E0/E20/E25/E30 in a ratio (edit blue shares; each row must sum to 100%)"
mx["A1"].font = TITLE
mx["A2"] = "Same distance driven across rows. OMC petrol income = blend-weighted throughput × petrol margin."; mx["A2"].font = NOTE
hdr = ["Mix scenario", "E0", "E20", "E25", "E30", "Σ", "Wtd drop", "Blend vol (bn L)",
       "Ethanol (bn L)", "Petrol MS (bn L)", "OMC petrol income (Rs cr)", "Extra vs all-E0 (Rs cr)"]
mx.append([]); mx.append(hdr)
for c in mx[4]:
    c.font = HEAD; c.fill = HFILL
p20, p25, p30 = R["E20 blend %"], R["E25 blend %"], R["E30 blend %"]
d20, d25, d30 = R["E20 mileage drop"], R["E25 mileage drop"], R["E30 mileage drop"]
mix_rows = [("S0 Today (~E20)", 0.05, 0.95, 0.00, 0.00),
            ("S1 E20 universal", 0.00, 1.00, 0.00, 0.00),
            ("S2 Transition FY27", 0.05, 0.55, 0.35, 0.05),
            ("S3 E25 majority FY28", 0.05, 0.20, 0.60, 0.15),
            ("S4 E30 push FY30", 0.05, 0.10, 0.25, 0.60)]
for i, (name, e0, e20, e25, e30) in enumerate(mix_rows):
    r = 5 + i
    mx.cell(r, 1, name)
    mx.cell(r, 2, e0).font = BLUE;  mx.cell(r, 2).number_format = PCT
    mx.cell(r, 3, e20).font = BLUE; mx.cell(r, 3).number_format = PCT
    mx.cell(r, 4, e25).font = BLUE; mx.cell(r, 4).number_format = PCT
    mx.cell(r, 5, e30).font = BLUE; mx.cell(r, 5).number_format = PCT
    mx.cell(r, 6, f"=SUM(B{r}:E{r})").number_format = PCT
    mx.cell(r, 7, f"=C{r}*{d20}+D{r}*{d25}+E{r}*{d30}").number_format = PCT     # E0 drop=0
    blend = f"({L0})*(B{r}+C{r}/(1-{d20})+D{r}/(1-{d25})+E{r}/(1-{d30}))"
    mx.cell(r, 8, f"={blend}").number_format = RS2
    mx.cell(r, 9, f"=({L0})*(C{r}/(1-{d20})*{p20}+D{r}/(1-{d25})*{p25}+E{r}/(1-{d30})*{p30})").number_format = RS2
    mx.cell(r, 10, f"=H{r}-I{r}").number_format = RS2
    mx.cell(r, 11, f"=H{r}*1000000000*{MGN_MS}/10000000").number_format = RS
    mx.cell(r, 12, f"=(H{r}-({L0}))*1000000000*{MGN_MS}/10000000").number_format = RS
mx.column_dimensions["A"].width = 22
for col in "BCDEFG": mx.column_dimensions[col].width = 10
for col in "HIJKL": mx.column_dimensions[col].width = 16
mx["A11"] = ("CBG caveat: this mileage-drop → extra-throughput effect is ETHANOL-ONLY. CBG cascaded "
             "into CNG must meet IS 16087 (min ~90% methane), so its calorific value matches fossil CNG "
             "(IS 15958) — no mileage loss, hence no analogous throughput uplift for the CNG book.")
mx["A11"].font = NOTE; mx["A11"].alignment = Alignment(wrap_text=True)
mx.merge_cells("A11:L13")

# ── Projection ──────────────────────────────────────────────────────────────
p = wb.create_sheet("Projection"); p.sheet_view.showGridLines = False
p["A1"] = "YoY projection — vehicle demand growth + ethanol blend roadmap"; p["A1"].font = TITLE
hdr = ["FY", "Blend drop", "Outlets", "Petrol blend (bn L)", "Diesel (bn L)",
       "OMC income (Rs cr)", "YoY Δ (Rs cr)", "Ethanol-attrib (Rs cr)"]
p.append([]); p.append(hdr)
for c in p[3]:
    c.font = HEAD; c.fill = HFILL
# roadmap: which drop applies each FY (references Inputs)
road = [("FY25-26", R["E20 mileage drop"]), ("FY26-27", R["E20 mileage drop"]),
        ("FY27-28", R["E25 mileage drop"]), ("FY28-29", R["E25 mileage drop"]),
        ("FY29-30", R["E30 mileage drop"])]
# L0 base and diesel base for year 0
for i, (fy, drop) in enumerate(road):
    r = 4 + i
    n = i  # years of growth applied
    L0_y = f"({L0})*(1+{GMS})^{n}"
    hsd_y = f"({bnL(HSD_MMT,DHSD)})*(1+{GHSD})^{n}"
    nro_y = f"{NRO}*(1+{GRO})^{n}"
    blendv = f"({L0_y}/(1-{drop}))"
    p.cell(r, 1, fy)
    p.cell(r, 2, f"={drop}").number_format = PCT
    p.cell(r, 3, f"={nro_y}").number_format = RS
    p.cell(r, 4, f"={blendv}").number_format = RS2
    p.cell(r, 5, f"={hsd_y}").number_format = RS2
    p.cell(r, 6, f"=D{r}*1000000000*{MGN_MS}/10000000+E{r}*1000000000*{MGN_HSD}/10000000").number_format = RS
    p.cell(r, 7, "" if i == 0 else f"=F{r}-F{r-1}").number_format = RS
    p.cell(r, 8, f"=(D{r}-({L0_y}))*1000000000*{MGN_MS}/10000000").number_format = RS
p.column_dimensions["A"].width = 10
for col in "BCDEFGH": p.column_dimensions[col].width = 16

# ── Checks ──────────────────────────────────────────────────────────────────
ck = wb.create_sheet("Checks"); ck.sheet_view.showGridLines = False
ck["A1"] = "Consistency checks (all must read PASS)"; ck["A1"].font = TITLE
ck.append([]); ck.append(["Check", "Result"])
for c in ck[3]:
    c.font = HEAD; c.fill = HFILL
checks = [
    ("Scenario E20 blend vol = base petrol volume", f'=IF(ABS(Scenarios!D5-{bnL(MS_MMT,DMS)})<0.01,"PASS","FAIL")'),
    ("Extra OMC income rises with blend (E20<E25<E30)", '=IF(AND(Scenarios!H5<Scenarios!H6,Scenarios!H6<Scenarios!H7),"PASS","FAIL")'),
    ("Pure petrol falls as blend rises (E20>E25>E30)", '=IF(AND(Scenarios!E5>Scenarios!E6,Scenarios!E6>Scenarios!E7),"PASS","FAIL")'),
    ("Projection FY25-26 income = Base total", f'=IF(ABS(Projection!F4-{BASE_TOT})<1,"PASS","FAIL")'),
    ("Projection income strictly increasing", '=IF(AND(Projection!F5>Projection!F4,Projection!F8>Projection!F7),"PASS","FAIL")'),
    ("Outlets grow YoY", '=IF(Projection!C8>Projection!C4,"PASS","FAIL")'),
    ("Every blend-mix row sums to 100%", '=IF(AND(BlendMix!F5=1,BlendMix!F6=1,BlendMix!F7=1,BlendMix!F8=1,BlendMix!F9=1),"PASS","FAIL")'),
    ("Mix OMC petrol income rises as pool shifts E0→E30", '=IF(AND(BlendMix!K6<BlendMix!K7,BlendMix!K7<BlendMix!K8,BlendMix!K8<BlendMix!K9),"PASS","FAIL")'),
    ("Mix pure-petrol falls as ethanol share rises", '=IF(BlendMix!J9<BlendMix!J6,"PASS","FAIL")'),
]
for i, (name, f) in enumerate(checks, start=4):
    ck.cell(i, 1, name); ck.cell(i, 2, f)
ck.column_dimensions["A"].width = 48; ck.column_dimensions["B"].width = 10

wb.save(OUT)
print(f"wrote {OUT}")
